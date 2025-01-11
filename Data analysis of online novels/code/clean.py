# 数据清洗
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('new_data.xlsx')
sheet = wb['Sheet1']
column_letter = 'G'
# 积分转换为数字格式
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=6, max_col=7):
    cell = row[0]
    if isinstance(cell.value, str):
        try:
            cell.value = int(cell.value.replace(',', ''))
        except ValueError:
            print(f"无法转换单元格 {cell.coordinate} 的值为数字格式")
wb.save('new_data.xlsx')

df = pd.read_excel('new_data.xlsx')
df['时间'] = pd.to_datetime(df['时间']).dt.strftime('%Y-%m-%d')
df['字数'] = df['字数'].astype(int)

df.to_excel('new_data.xlsx', index=False)